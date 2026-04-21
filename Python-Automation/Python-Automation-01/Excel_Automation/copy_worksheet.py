from openpyxl import load_workbook

source = load_workbook("jLoka_01.xlsx")
destination = load_workbook("jloka_02.xlsx")

source_sheet = source.worksheets[2]
destination_sheet = destination.active

max_row = source_sheet.max_row

for row in range(1, max_row + 1):
    for column in range(1, 3):
        c = source_sheet.cell(row=row, column=column)
        destination_sheet.cell(row=row, column=column).value = c.value

destination.save("jloka_02.xlsx")
destination.close()