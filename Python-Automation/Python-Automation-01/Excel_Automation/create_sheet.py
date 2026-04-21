from openpyxl import Workbook

from datetime import datetime

workbook = Workbook()

workbook.create_sheet("jloka_01")
workbook.create_sheet("jloka_02", index=0)

sheet = workbook['jloka_01']
sheet["A1"] = datetime.now()
sheet["B1"] = datetime.now()
workbook.copy_worksheet(sheet)
print("The sheet names are: ", workbook.sheetnames)

workbook.save("jLoka_01.xlsx")
workbook.close()